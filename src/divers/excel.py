# -*- coding: utf-8 -*-
"""
Functions to interact with excel files

- autofit_column_widths
    autofits the column widths in an excel file and centers the columns

- delete_file
    sends a warning if a file already exist and deletes the file is user
    responds 'y' to 'Do you want to overwrite he file?'

- append_df_to_excel
    appends a DataFrame to existing Excel file [filename] into [sheet_name]
    Sheet.
    If [filename] doesn't exist, then this function will create it.
"""
__version__ = '1.0'
__author__ = 'Noemie Fedon'

import sys
import os
import pathlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def autofit_column_widths(filename):
    """
    autofits the column widths in an excel file
    """
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = load_workbook(filename)
    sheetnames = writer.book.sheetnames

#    for worksheet in writer.book[sheetnames]: # NO !

    for ind_sheet in range(len(sheetnames)):
        worksheet = writer.book[sheetnames[ind_sheet]]

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            # Since Openpyxl 2.6, the column name is  ".column_letter"
            # as .column became the column number (1-based)
            for cell in col:
                try: # Necessary to avoid error on empty
                    cell.alignment = Alignment(horizontal='center')
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    writer.save()
    return 0


def delete_file(filename):
    """
    sends a warning if a file already exist and deletes the file is user
    responds 'y' to 'Do you want to overwrite he file?'
    """
    path = pathlib.Path(filename)
    if path.exists():
        message = 'Do you want to overwrite the file ' \
        + filename + '?   type \'y\' for yes or \'n\' for no\n'
        answer = input(message)
        answer = str(answer)
        if answer in {'y', 'Y', 'Yes', 'yes', 'YES'}:
            os.remove(filename)
        else:
            sys.exit('The algorithm is interrupted.')

    return 0


def append_df_to_excel(filename, dataf, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [dataf] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      dataf : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing dataf and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    dataf.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()

if __name__ == "__main__":
    print('When two pages are saved successively, the first page is lost')
    writer = pd.ExcelWriter('test.xlsx')
    table_const = pd.DataFrame()
    i = 1
    table_const.loc[i, 'When?'] = 'Now'
    table_const.loc[i, 'Where?'] = 'There'
    i = 2
    table_const.loc[i, 'When?'] = 'Later'
    table_const.loc[i, 'Where?'] = 'Here'
    table_const.to_excel(writer, 'Best design' + str(0))
    writer.save()
    table_const = pd.DataFrame()
    i = 1
    table_const.loc[i, 'When?'] = 'Now2'
    table_const.loc[i, 'Where?'] = 'There2'
    i = 2
    table_const.loc[i, 'When?'] = 'Later2'
    table_const.loc[i, 'Where?'] = 'Here2'
    table_const.to_excel(writer, 'Best design' + str(2))
    writer.save()

    print('When two pages are saved successively, the first page is not lost')
    filename = 'test.xlsx'
    writer = pd.ExcelWriter(filename)
    table_const = pd.DataFrame()
    i = 1
    table_const.loc[i, 'When?'] = 'Now'
    table_const.loc[i, 'Where?'] = 'There'
    i = 2
    table_const.loc[i, 'When?'] = 'Later'
    table_const.loc[i, 'Where?'] = 'Here'
    append_df_to_excel(filename, table_const, 'Best' + str(0))
    table_const = pd.DataFrame()
    i = 1
    table_const.loc[i, 'When?'] = 'Now2'
    table_const.loc[i, 'Where?'] = 'There2'
    i = 2
    table_const.loc[i, 'When?'] = 'Later2'
    table_const.loc[i, 'Where?'] = 'Here2'
    append_df_to_excel(filename, table_const, 'Best' + str(1))
